#!/usr/bin/env python3
"""Convert the published Taufik et al. historical groundwater workbook to JagaRawa CSV."""

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def station_metadata(path: Path) -> dict[str, dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        rows = {row["station_id"]: row for row in csv.DictReader(handle)}
    incomplete = [station for station, row in rows.items() if not row.get("latitude") or not row.get("longitude")]
    if incomplete:
        raise ValueError(f"Missing verified coordinates for: {', '.join(incomplete)}")
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=Path("data/raw/taufik-2022-groundwater.xlsx"))
    parser.add_argument("--stations", type=Path, default=Path("data/templates/taufik-2022-stations.csv"))
    parser.add_argument("--output", type=Path, default=Path("data/processed/taufik-2022-water-table.csv"))
    args = parser.parse_args()
    stations = station_metadata(args.stations)
    workbook = load_workbook(args.input, data_only=True, read_only=True)
    sheet = workbook["gwt"]
    args.output.parent.mkdir(parents=True, exist_ok=True)

    with args.output.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["station_id", "latitude", "longitude", "observed_at", "water_table_depth_m", "unit", "source", "quality_flag", "peatland"])
        writer.writeheader()
        for date, groundwater_m, station_id, peatland in sheet.iter_rows(min_row=2, values_only=True):
            if not date or groundwater_m is None or station_id not in stations:
                continue
            # The source workbook records below-ground measurements as negative metres.
            writer.writerow({
                "station_id": station_id,
                "latitude": stations[station_id]["latitude"],
                "longitude": stations[station_id]["longitude"],
                "observed_at": f"{date.isoformat()}T00:00:00Z",
                "water_table_depth_m": f"{-float(groundwater_m):.3f}",
                "unit": "m",
                "source": "Taufik et al. 2022 Data in Brief, historical research dataset",
                "quality_flag": "historical_research",
                "peatland": peatland,
            })


if __name__ == "__main__":
    main()
